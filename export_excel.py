from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io

def format_name(full_name):
    """Перетворити 'Прізвище Ім'я По-батькові' на 'Ім'я ПРІЗВИЩЕ'"""
    parts = full_name.split()
    if len(parts) >= 2:
        surname = parts[0].upper()  # Прізвище великими літерами
        first_name = parts[1]  # Ім'я
        return f"{first_name} {surname}"
    return full_name  # Якщо формат незрозумілий, повернути як є

def create_teacher_report_excel(data, year, class_name, teacher, subject, semester):
    """Створити Excel звіт для вчителя"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name}_{subject[:20]}"
    
    # Заголовок - об'єднати до N колонки
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f'Коломийський ліцей "Коломийська гімназія імені Михайла Грушевського"'
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:N2')
    subtitle_cell = ws['A2']
    semester_display = 'I' if str(semester) == '1' or semester == 1 else 'II'
    subtitle_cell.value = f'Звіт вчителя за {semester_display} семестр {year} н.р.'
    subtitle_cell.font = Font(size=12, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:N3')
    info_cell = ws['A3']
    info_cell.value = f'Клас: {class_name} | Вчитель: {teacher} | Предмет: {subject}'
    info_cell.font = Font(size=11)
    info_cell.alignment = Alignment(horizontal='center')
    
    current_row = 4
    if subject == 'Фізична культура' and data.get('pe_exempted_count', 0) > 0:
        ws.merge_cells(f'A{current_row}:N{current_row}')
        info_exempted = ws[f'A{current_row}']
        total_students = data['student_count'] + data['pe_exempted_count']
        info_exempted.value = f'ℹ️ Всього учнів у класі: {total_students} | Звільнені від занять: {data["pe_exempted_count"]} | Підлягають оцінюванню: {data["student_count"]}'
        info_exempted.font = Font(size=10, italic=True, color='0066CC')
        info_exempted.alignment = Alignment(horizontal='center')
        current_row += 1
    
    # Заголовки колонок
    headers = ['Учні', 'н/а', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
    
    current_row += 1
    header_row = current_row
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Дані оцінок
    student_count = data.get('student_count', 0)
    grades = data.get('grades', {})
    
    ws.cell(row=header_row, column=1, value='Кількість учнів класу').font = Font(bold=True)
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='left')
    
    current_row = header_row + 1
    
    # Показати кількість учнів та розподіл оцінок
    ws.cell(row=current_row, column=1, value=student_count).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    
    not_assessed = student_count
    for i in range(1, 13):
        grade_count = int(grades.get(f'grade{i}', 0))
        not_assessed -= grade_count
        ws.cell(row=current_row, column=i+2, value=grade_count).alignment = Alignment(horizontal='center')
    
    ws.cell(row=current_row, column=2, value=not_assessed).alignment = Alignment(horizontal='center')
    
    # Статистика
    stats = data.get('statistics', {})
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1, value='СТАТИСТИЧНІ ПОКАЗНИКИ').font = Font(bold=True, size=11)
    current_row += 1
    
    stats_data = [
        ('Середній бал (СБ):', stats.get('avgScore', '0')),
        ('Ступінь навченості (СН):', stats.get('learningLevel', '0%')),
        ('Коефіцієнт якості знань (КЯЗ):', stats.get('qualityCoeff', '0%')),
        ('Якість знань (ЯЗ):', stats.get('qualityPercent', '0%')),
        ('Коефіцієнт результативності (КР):', stats.get('resultCoeff', '0%'))
    ]
    
    for label, value in stats_data:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value).font = Font(size=11, color='0000FF')
        current_row += 1
    
    # Місце для підписів
    current_row += 3
    
    ws.cell(row=current_row, column=1, value='Директор')
    ws.cell(row=current_row, column=6, value='________________')
    ws.cell(row=current_row, column=10, value='Володимир ТКАЧУК')
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.merge_cells(f'F{current_row}:I{current_row}')
    ws.merge_cells(f'J{current_row}:N{current_row}')
    
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'A{current_row}'].font = Font(size=11)
    ws[f'F{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'J{current_row}'].alignment = Alignment(horizontal='left')
    ws[f'J{current_row}'].font = Font(size=11)
    
    current_row += 3
    ws.cell(row=current_row, column=1, value='Вчитель')
    ws.cell(row=current_row, column=6, value='________________')
    ws.cell(row=current_row, column=10, value=format_name(teacher))
    
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.merge_cells(f'F{current_row}:I{current_row}')
    ws.merge_cells(f'J{current_row}:N{current_row}')
    
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
    ws[f'A{current_row}'].font = Font(size=11)
    ws[f'F{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'J{current_row}'].alignment = Alignment(horizontal='left')
    ws[f'J{current_row}'].font = Font(size=11)
    
    column_widths = {'A': 35, 'B': 8, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 
                     'H': 8, 'I': 8, 'J': 8, 'K': 8, 'L': 8, 'M': 8, 'N': 8}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def create_class_report_excel(class_data, class_name, year, semester, class_head_name=None):
    """Створити Excel звіт по класу з об'єднанням підгруп"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name}"
    
    # ✅ КРОК 1: СПОЧАТКУ створити merged_data
    filled_data = [item for item in class_data if item.get('filled', False)]
    
    # Групувати дані по предметах
    subject_groups = {}
    for item in filled_data:
        subject = item['subject']
        if subject not in subject_groups:
            subject_groups[subject] = []
        subject_groups[subject].append(item)
    
    # Об'єднати дані для кожного предмету
    merged_data = []
    for subject, items in sorted(subject_groups.items()):
        if len(items) == 1:
            merged_data.append(items[0])
        else:
            # Кілька вчителів (підгрупи) - об'єднуємо
            teachers = [item['teacher'] for item in items]
            teacher_display = ' / '.join(teachers)
            
            total_students = sum(item['student_count'] for item in items)
            
            merged_grades = {}
            for grade_key in ['grade1', 'grade2', 'grade3', 'grade4', 'grade5', 'grade6',
                            'grade7', 'grade8', 'grade9', 'grade10', 'grade11', 'grade12', 'gradeNA']:
                merged_grades[grade_key] = sum(int(item['grades'].get(grade_key, 0)) for item in items)
            
            avg_score = sum(float(item['statistics']['avgScore']) for item in items) / len(items)
            avg_learning = sum(float(item['statistics']['learningLevel'].replace('%', '')) for item in items) / len(items)
            avg_quality_coeff = sum(float(item['statistics']['qualityCoeff'].replace('%', '')) for item in items) / len(items)
            avg_quality_percent = sum(float(item['statistics']['qualityPercent'].replace('%', '')) for item in items) / len(items)
            avg_result = sum(float(item['statistics']['resultCoeff'].replace('%', '')) for item in items) / len(items)
            
            merged_item = {
                'subject': subject,
                'teacher': teacher_display,
                'student_count': total_students,
                'grades': merged_grades,
                'statistics': {
                    'avgScore': f'{avg_score:.2f}',
                    'learningLevel': f'{avg_learning:.2f}%',
                    'qualityCoeff': f'{avg_quality_coeff:.2f}%',
                    'qualityPercent': f'{avg_quality_percent:.2f}%',
                    'resultCoeff': f'{avg_result:.2f}%'
                },
                'filled': True
            }
            merged_data.append(merged_item)
    
    # ✅ ВИПРАВЛЕНО: Шапка звіту - всі рядки на всю ширину по центру

    # Рядок 1: Назва ліцею (по центру)
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = f'Коломийський ліцей "Коломийська гімназія імені Михайла Грушевського"'
    title_cell.font = Font(size=12, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Рядок 2: Попредметний звіт (по центру)
    ws.merge_cells('A2:O2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f'Попредметний звіт {class_name} класу'
    subtitle_cell.font = Font(size=11, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Рядок 3: Інформація про клас (по центру)
    if merged_data:
        student_count = max(item.get('student_count', 0) for item in merged_data)
        
        ws.merge_cells('A3:O3')
        info_cell = ws['A3']
        info_cell.value = f'Кількість учнів у класі: {student_count}  |  Навчальний рік: {year}  |  Семестр: {semester}'
        info_cell.font = Font(size=11, bold=True, color='0066CC')
        info_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Рядок 4: Порожній
    ws.row_dimensions[4].height = 5

    # Рядок 5: Заголовки таблиці
    headers = ['№', 'Предмет', 'Вчитель', 'Звільн.', 'н/а', 'початковий', 'середній', 'достатній', 
            'високий', 'СБ', 'СН(%)', 'СН', 'КЯЗ', 'ЯЗ', 'КР']

    header_row = 5
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Дані
    row = header_row + 1
    
    # Змінні для обчислення зведеної статистики
    total_stats = {
        'not_assessed': [], 'initial': [], 'average': [], 'sufficient': [], 'high': [],
        'not_assessed_count': 0, 'initial_count': 0, 'average_count': 0, 
        'sufficient_count': 0, 'high_count': 0,
        'avg_scores': [], 'quality_coeffs': [], 'result_coeffs': []
    }
    
    # ✅ ЗМІНЕНО: Використовуємо merged_data
    for idx, item in enumerate(merged_data, start=1):
        stats = item['statistics']
        grades = item['grades']
        student_count = item['student_count']
        
        # Розрахунок
        g3 = int(grades.get('grade3', 0))
        g2 = int(grades.get('grade2', 0))
        g1 = int(grades.get('grade1', 0))
        initial = g3 + g2 + g1
        initial_pct = (initial / student_count * 100) if student_count > 0 else 0
        
        g6 = int(grades.get('grade6', 0))
        g5 = int(grades.get('grade5', 0))
        g4 = int(grades.get('grade4', 0))
        average = g6 + g5 + g4
        average_pct = (average / student_count * 100) if student_count > 0 else 0
        
        g9 = int(grades.get('grade9', 0))
        g8 = int(grades.get('grade8', 0))
        g7 = int(grades.get('grade7', 0))
        sufficient = g9 + g8 + g7
        sufficient_pct = (sufficient / student_count * 100) if student_count > 0 else 0
        
        g12 = int(grades.get('grade12', 0))
        g11 = int(grades.get('grade11', 0))
        g10 = int(grades.get('grade10', 0))
        high = g12 + g11 + g10
        high_pct = (high / student_count * 100) if student_count > 0 else 0
        
        total_graded = initial + average + sufficient + high
        not_assessed = max(0, student_count - total_graded)
        not_assessed_pct = (not_assessed / student_count * 100) if student_count > 0 else 0
        
        # Накопичення для зведеної статистики
        total_stats['not_assessed'].append(not_assessed_pct)
        total_stats['initial'].append(initial_pct)
        total_stats['average'].append(average_pct)
        total_stats['sufficient'].append(sufficient_pct)
        total_stats['high'].append(high_pct)
        total_stats['not_assessed_count'] += not_assessed
        total_stats['initial_count'] += initial
        total_stats['average_count'] += average
        total_stats['sufficient_count'] += sufficient
        total_stats['high_count'] += high
        total_stats['avg_scores'].append(float(stats['avgScore']))
        total_stats['quality_coeffs'].append(float(stats['qualityCoeff'].replace('%', '')))
        total_stats['result_coeffs'].append(float(stats['resultCoeff'].replace('%', '')))
        
        # СН текст
        learning_level_num = float(stats['learningLevel'].replace('%', ''))
        if learning_level_num >= 64:
            sn_text = 'високий ступінь навченості'
        elif learning_level_num >= 36:
            sn_text = 'достатній ступінь навченості'
        else:
            sn_text = 'середній ступінь навченості'
        
        # Звільнені
        exempted_display = '-'
        if item.get('subject') == 'Фізична культура' and item.get('pe_exempted_count', 0) > 0:
            exempted_display = str(item['pe_exempted_count'])
        
        # ✅ ДОДАНО: Стовпець "Вчитель"
        row_data = [
            idx,
            item['subject'],
            item.get('teacher', '-'),  # ✅ НОВИЙ СТОВПЕЦЬ
            exempted_display,
            f"{not_assessed_pct:.2f}% ({not_assessed})",
            f"{initial_pct:.2f}% ({initial})",
            f"{average_pct:.2f}% ({average})",
            f"{sufficient_pct:.2f}% ({sufficient})",
            f"{high_pct:.2f}% ({high})",
            stats['avgScore'],
            stats['learningLevel'],
            sn_text,
            stats['qualityCoeff'],
            stats['qualityPercent'],
            stats['resultCoeff']
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            # Вирівнювання
            if col == 2 or col == 3:  # Предмет та Вчитель - ліворуч
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ✅ ДОДАНО: Перенесення тексту для стовпця "Вчитель"
            if col == 3:  # Стовпець "Вчитель"
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # ✅ ДОДАНО: Встановити висоту рядка
        ws.row_dimensions[row].height = 30
        
        row += 1

        # ✅ ДОДАНО: Зведена статистика класу
    if total_stats['avg_scores']:
        row += 2
        ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО: N → O
        stats_header = ws.cell(row=row, column=1, value='ЗВЕДЕНІ ПОКАЗНИКИ КЛАСУ')
        stats_header.font = Font(size=12, bold=True, color='FFFFFF')
        stats_header.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        stats_header.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Обчислити середні значення
        avg_score = sum(total_stats['avg_scores']) / len(total_stats['avg_scores'])
        avg_quality = sum(total_stats['quality_coeffs']) / len(total_stats['quality_coeffs'])
        avg_result = sum(total_stats['result_coeffs']) / len(total_stats['result_coeffs'])
        
        # Обчислити середній рівень навченості
        avg_sn = (avg_quality + avg_result) / 2
        
        summary_data = [
            ('Середній бал по класу:', f'{avg_score:.2f}'),
            ('Коефіцієнт якості знань (КЯЗ):', f'{avg_quality:.2f}%'),
            ('Коефіцієнт результативності (КР):', f'{avg_result:.2f}%'),
            ('Рівень навченості (СН):', f'{avg_sn:.2f}%'),
        ]
        
        for label, value in summary_data:
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
            
            ws.merge_cells(f'H{row}:O{row}')  # ✅ ЗМІНЕНО: N → O
            ws.cell(row=row, column=8, value=value).font = Font(size=11, color='0000FF', bold=True)
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
            
            row += 1
        
        row += 1
        ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО
        distribution_header = ws.cell(row=row, column=1, value='Розподіл за рівнями:')
        distribution_header.font = Font(size=11, bold=True)
        distribution_header.alignment = Alignment(horizontal='left')
        
        row += 1
        
        # Розподіл
        num_subjects = len(total_stats['avg_scores'])
        distribution_data = [
            ('Високий рівень:', f'{total_stats["high_count"]} результатів ({sum(total_stats["high"])/num_subjects:.2f}%)'),
            ('Достатній рівень:', f'{total_stats["sufficient_count"]} результатів ({sum(total_stats["sufficient"])/num_subjects:.2f}%)'),
            ('Середній рівень:', f'{total_stats["average_count"]} результатів ({sum(total_stats["average"])/num_subjects:.2f}%)'),
            ('Початковий рівень:', f'{total_stats["initial_count"]} результатів ({sum(total_stats["initial"])/num_subjects:.2f}%)'),
            ('Не оцінено:', f'{total_stats["not_assessed_count"]} результатів ({sum(total_stats["not_assessed"])/num_subjects:.2f}%)'),
        ]
        
        for label, value in distribution_data:
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value=label).font = Font(size=10)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2)
            
            ws.merge_cells(f'H{row}:O{row}')  # ✅ ЗМІНЕНО
            ws.cell(row=row, column=8, value=value).font = Font(size=10)
            ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # ✅ ДОДАНО: Топ-3 та проблемні предмети
        row += 1
        
        # Сортуємо предмети по середньому балу
        subject_scores = [(item['subject'], float(item['statistics']['avgScore'])) 
                         for item in merged_data]
        subject_scores_sorted = sorted(subject_scores, key=lambda x: x[1], reverse=True)
        
        # Топ-3
        ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО
        top_header = ws.cell(row=row, column=1, value='Найвищі показники (за середнім балом):')
        top_header.font = Font(size=11, bold=True, color='006600')
        top_header.alignment = Alignment(horizontal='left')
        row += 1
        
        for i, (subject, score) in enumerate(subject_scores_sorted[:3], start=1):
            ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО
            ws.cell(row=row, column=1, value=f'{i}. {subject} — СБ: {score:.2f}').font = Font(size=10)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2)
            row += 1
        
        # Проблемні (якщо є предмети з СБ < 7)
        problem_subjects = [(subj, score) for subj, score in subject_scores_sorted if score < 7]
        
        if problem_subjects:
            row += 1
            ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО
            problem_header = ws.cell(row=row, column=1, value='Потребують уваги (СБ < 7):')
            problem_header.font = Font(size=11, bold=True, color='CC0000')
            problem_header.alignment = Alignment(horizontal='left')
            row += 1
            
            for i, (subject, score) in enumerate(problem_subjects[:3], start=1):
                ws.merge_cells(f'A{row}:O{row}')  # ✅ ЗМІНЕНО
                ws.cell(row=row, column=1, value=f'{i}. {subject} — СБ: {score:.2f}').font = Font(size=10, color='CC0000')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2)
                row += 1
    
    # ✅ ОНОВЛЕНО: Дата та підписи
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    date_cell = ws.cell(row=row, column=1, value=f'Дата складання звіту: "___" __________ 202__ р.')
    date_cell.font = Font(size=10)
    date_cell.alignment = Alignment(horizontal='left')
    
    # Класний керівник
    row += 2
    ws.merge_cells(f'H{row}:J{row}')
    ws.cell(row=row, column=8, value='Класний керівник')
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=8).font = Font(size=11)
    
    ws.merge_cells(f'K{row}:O{row}')  # ✅ ЗМІНЕНО: N → O
    if class_head_name:
        class_head_formatted = format_name(class_head_name)
        ws.cell(row=row, column=11, value=class_head_formatted)
    else:
        ws.cell(row=row, column=11, value='___________________________')
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=11).font = Font(size=11)
    
    # /підпис/
    row += 1
    ws.merge_cells(f'I{row}:J{row}')
    ws.cell(row=row, column=9, value='/підпис/')
    ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=9).font = Font(size=10, italic=True)
    
    # Директор
    row += 2
    ws.merge_cells(f'H{row}:J{row}')
    ws.cell(row=row, column=8, value='Директор')
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=8).font = Font(size=11)
    
    ws.merge_cells(f'K{row}:O{row}')  # ✅ ЗМІНЕНО: N → O
    ws.cell(row=row, column=11, value='Володимир ТКАЧУК')
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=11).font = Font(size=11)
    
    # /підпис/
    row += 1
    ws.merge_cells(f'I{row}:J{row}')
    ws.cell(row=row, column=9, value='/підпис/')
    ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=9).font = Font(size=10, italic=True)
    
    # ✅ ОНОВЛЕНО: Ширина колонок (додано стовпець O для вчителя)
    column_widths = {'A': 5, 'B': 30, 'C': 25, 'D': 8, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 14,
                     'J': 8, 'K': 10, 'L': 30, 'M': 8, 'N': 8, 'O': 8}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def create_school_report_excel(school_data, all_monitoring, year, semester):
    """Створити Excel звіт по школі з усіма показниками"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт по школі"
    
    # Заголовок
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = 'Звіт по класах'
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:M2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f'за {year} навчальний рік'
    subtitle_cell.font = Font(size=12, bold=True)
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Заголовки колонок
    headers = ['Клас', 'н/а', 'початковий', 'середній', 'достатній', 
               'високий', 'СБ', 'СН(%)', 'СН', 'КЯЗ', 'ЯЗ', 'КР']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Підготовка даних по класах
    class_stats = {}
    
    for record in all_monitoring:
        class_name = record['class']
        if class_name not in class_stats:
            class_stats[class_name] = {
                'subjects': [],
                'total_students': 0
            }
        
        class_stats[class_name]['subjects'].append(record)
        current_count = record['student_count']
        if current_count > class_stats[class_name]['total_students']:
            class_stats[class_name]['total_students'] = current_count
    
    # Дані по класах
    row = 5
    
    overall_stats = {
        'not_assessed': [],
        'initial': [],
        'average': [],
        'sufficient': [],
        'high': [],
        # ✅ ДОДАНО: Абсолютні числа
        'not_assessed_count': [],
        'initial_count': [],
        'average_count': [],
        'sufficient_count': [],
        'high_count': [],
        'avg_score': [],
        'learning_level': [],
        'quality_coeff': [],
        'quality_percent': [],
        'result_coeff': []
    }
    
    for class_name in sorted(class_stats.keys()):
        class_data = class_stats[class_name]
        subjects = class_data['subjects']
        student_count = class_data['total_students']
        
        if not subjects or student_count == 0:
            continue
        
        # Розрахунок середніх для класу
        total_not_assessed = 0
        total_initial = 0
        total_average = 0
        total_sufficient = 0
        total_high = 0
        # ✅ ДОДАНО: Лічильники абсолютних чисел
        sum_not_assessed_count = 0
        sum_initial_count = 0
        sum_average_count = 0
        sum_sufficient_count = 0
        sum_high_count = 0
        sum_avg_score = 0
        sum_learning_level = 0
        sum_quality_coeff = 0
        sum_quality_percent = 0
        sum_result_coeff = 0
        
        for subject in subjects:
            grades = subject['grades']
            subject_student_count = subject['student_count']
            
            if subject_student_count == 0:
                continue
            
            # Рівні
            g3 = int(grades.get('grade3', 0))
            g2 = int(grades.get('grade2', 0))
            g1 = int(grades.get('grade1', 0))
            initial = g3 + g2 + g1
            
            g6 = int(grades.get('grade6', 0))
            g5 = int(grades.get('grade5', 0))
            g4 = int(grades.get('grade4', 0))
            average = g6 + g5 + g4
            
            g9 = int(grades.get('grade9', 0))
            g8 = int(grades.get('grade8', 0))
            g7 = int(grades.get('grade7', 0))
            sufficient = g9 + g8 + g7
            
            g12 = int(grades.get('grade12', 0))
            g11 = int(grades.get('grade11', 0))
            g10 = int(grades.get('grade10', 0))
            high = g12 + g11 + g10
            
            total_graded = initial + average + sufficient + high
            not_assessed = max(0, subject_student_count - total_graded)
            
            # ✅ ДОДАНО: Накопичення абсолютних чисел
            sum_not_assessed_count += not_assessed
            sum_initial_count += initial
            sum_average_count += average
            sum_sufficient_count += sufficient
            sum_high_count += high
            
            total_not_assessed += (not_assessed / subject_student_count * 100)
            total_initial += (initial / subject_student_count * 100)
            total_average += (average / subject_student_count * 100)
            total_sufficient += (sufficient / subject_student_count * 100)
            total_high += (high / subject_student_count * 100)
            
            # Статистика
            stats = subject['statistics']
            sum_avg_score += float(stats.get('avgScore', 0))
            sum_learning_level += float(stats.get('learningLevel', '0%').replace('%', ''))
            sum_quality_coeff += float(stats.get('qualityCoeff', '0%').replace('%', ''))
            sum_quality_percent += float(stats.get('qualityPercent', '0%').replace('%', ''))
            sum_result_coeff += float(stats.get('resultCoeff', '0%').replace('%', ''))
        
        num_subjects = len(subjects)
        
        if num_subjects == 0:
            continue
        
        avg_not_assessed = total_not_assessed / num_subjects
        avg_initial = total_initial / num_subjects
        avg_average = total_average / num_subjects
        avg_sufficient = total_sufficient / num_subjects
        avg_high = total_high / num_subjects
        avg_score = sum_avg_score / num_subjects
        avg_learning_level = sum_learning_level / num_subjects
        avg_quality_coeff = sum_quality_coeff / num_subjects
        avg_quality_percent = sum_quality_percent / num_subjects
        avg_result_coeff = sum_result_coeff / num_subjects
        
        # СН текст
        if avg_learning_level >= 64:
            sn_text = 'високий ступінь навченості'
        elif avg_learning_level >= 36:
            sn_text = 'достатній ступінь навченості'
        else:
            sn_text = 'середній ступінь навченості'
        
        # ✅ ОНОВЛЕНО: Формат "60% (18)"
        row_data = [
            class_name,
            f"{avg_not_assessed:.2f}% ({sum_not_assessed_count})",
            f"{avg_initial:.2f}% ({sum_initial_count})",
            f"{avg_average:.2f}% ({sum_average_count})",
            f"{avg_sufficient:.2f}% ({sum_sufficient_count})",
            f"{avg_high:.2f}% ({sum_high_count})",
            f"{avg_score:.2f}",
            f"{avg_learning_level:.2f}%",
            sn_text,
            f"{avg_quality_coeff:.2f}%",
            f"{avg_quality_percent:.2f}%",
            f"{avg_result_coeff:.2f}%"
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Накопичення для загального
        overall_stats['not_assessed'].append(avg_not_assessed)
        overall_stats['initial'].append(avg_initial)
        overall_stats['average'].append(avg_average)
        overall_stats['sufficient'].append(avg_sufficient)
        overall_stats['high'].append(avg_high)
        # ✅ ДОДАНО
        overall_stats['not_assessed_count'].append(sum_not_assessed_count)
        overall_stats['initial_count'].append(sum_initial_count)
        overall_stats['average_count'].append(sum_average_count)
        overall_stats['sufficient_count'].append(sum_sufficient_count)
        overall_stats['high_count'].append(sum_high_count)
        overall_stats['avg_score'].append(avg_score)
        overall_stats['learning_level'].append(avg_learning_level)
        overall_stats['quality_coeff'].append(avg_quality_coeff)
        overall_stats['quality_percent'].append(avg_quality_percent)
        overall_stats['result_coeff'].append(avg_result_coeff)
        
        row += 1
    
    # ✅ ОНОВЛЕНО: Рядок "середні дані по ліцею"
    if overall_stats['avg_score']:
        avg_row = ws.cell(row=row, column=1, value='середні дані по ліцею')
        avg_row.font = Font(bold=True)
        avg_row.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        total_not_assessed_count = sum(overall_stats['not_assessed_count'])
        total_initial_count = sum(overall_stats['initial_count'])
        total_average_count = sum(overall_stats['average_count'])
        total_sufficient_count = sum(overall_stats['sufficient_count'])
        total_high_count = sum(overall_stats['high_count'])
        
        overall_data = [
            f"{sum(overall_stats['not_assessed'])/len(overall_stats['not_assessed']):.2f}% ({total_not_assessed_count})",
            f"{sum(overall_stats['initial'])/len(overall_stats['initial']):.2f}% ({total_initial_count})",
            f"{sum(overall_stats['average'])/len(overall_stats['average']):.2f}% ({total_average_count})",
            f"{sum(overall_stats['sufficient'])/len(overall_stats['sufficient']):.2f}% ({total_sufficient_count})",
            f"{sum(overall_stats['high'])/len(overall_stats['high']):.2f}% ({total_high_count})",
            f"{sum(overall_stats['avg_score'])/len(overall_stats['avg_score']):.2f}",
            f"{sum(overall_stats['learning_level'])/len(overall_stats['learning_level']):.2f}%",
            'високий ступінь навченості',
            f"{sum(overall_stats['quality_coeff'])/len(overall_stats['quality_coeff']):.2f}%",
            f"{sum(overall_stats['quality_percent'])/len(overall_stats['quality_percent']):.2f}%",
            f"{sum(overall_stats['result_coeff'])/len(overall_stats['result_coeff']):.2f}%"
        ]
        
        for col, value in enumerate(overall_data, start=2):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Підпис директора
    row += 2
    ws.merge_cells(f'A{row}:M{row}')
    ws.cell(row=row, column=1, value='Директор                          /підпис/              Володимир ТКАЧУК')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1
    ws.merge_cells(f'A{row}:M{row}')
    ws.cell(row=row, column=1, value='                  МП')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    
    # ✅ ОНОВЛЕНО: Збільшено ширину колонок для формату "60% (18)"
    column_widths = {'A': 20, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14,
                     'G': 8, 'H': 10, 'I': 30, 'J': 10, 'K': 10, 'L': 10, 'M': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output